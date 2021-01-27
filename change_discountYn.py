# Q. 카테고리별 상품설정
# S. 시간대별할인상품설정

import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('상품관리.xlsx')
ws = wb['상품관리']

# 현재시간에 따라 할인여부를 바꿔준다.
def change_discountYn():
    now = datetime.now()
    hour = now.hour # 현재 시간
    
    for rowNum in range(2, ws.max_row+1):
        disc_time = ws.cell(column=6, row=rowNum).value
        disc_Yn = ws.cell(column=4, row=rowNum).value
        
        # 12시 이전에 할인이 적용되는 상품의 경우
        if disc_time == '12시 이전':
            if now.hour < 12:   # 현재시간이 12시 이전이면 
                ws['D'+ str(rowNum)] = 'YES' # 할인여부를 Yes로 전환
            else:
                disc_Yn = 'NO'
                
        # 12시 이후에 할인이 적용되는 상품의 경우할인여부를 Yes로 전환           
        elif disc_time == '12시 이후':
            if now.hour < 12:   # 현재시간이 12시 이전이면
                disc_Yn = 'NO'  # 할인여부를 NO로 전환
            else:
                ws['D' + str(rowNum)] = 'YES'

    wb.save('상품관리.xlsx')
    
#관리자 할인시간/할인가격 설정
def discount_tp():
    while True:
        print('<상품종류>')
        print('1.물\n2.아메리카노\n3.코카콜라\n4.신라면\n5.삼양라면\n6.짜파게티')
        print('7.꼬깔콘\n8.포카칩\n9.새우깡\n')
        
        print('*원하는 기능의 번호를 선택해주세요.')
        number = int(input('0. 종료\n1. 할인시간 설정\n2. 할인가격 설정\n'))

        if( number == 0):
            return
        
        product = int(input('할인할 상품의 번호를 입력해주세요.'))
        if ( number == 1):
            discount_time = input('할인시간을 입력해주세요.(12시 이전,12시 이후) ')
            ws.cell(row= product + 1, column = 6).value = discount_time
        else:
            discount_price = int(input('할인하여 판매할 가격을 입력해주세요. '))

            if(discount_price > ws.cell(row = product+1,column=3).value):
                print('원금을 초과한 할인가격을 설정할 수 없습니다.')
                print('다시 입력해주세요.')
                continue
            ws.cell(row=product+1,column=5).value = discount_price
            
        wb.save('상품관리.xlsx')

change_discountYn()
discount_tp()
change_discountYn()
