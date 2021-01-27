# 이주연
# H. 좌석 관리 시뮬레이션 (+ O. 좌석속성설정 개념 차용-최영주)

import random
from openpyxl import load_workbook

wb = load_workbook("좌석관리.xlsx")
ws = wb.active

#난수 생성
def create_random():

    for i in range(2,31):
        rand_num = random.randint(0,1)
        ws.cell(row=i, column=3).value = bool(rand_num)
        wb.save('좌석관리.xlsx')
  
create_random() 

