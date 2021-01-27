# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'find_id.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *

libpaths = QApplication.libraryPaths() 
libpaths.append("C:\\Users\김민휘\AppData\Local\Programs\Python\Python36-32\Lib\site-packages\PyQt5\Qt\plugins") 
QApplication.setLibraryPaths(libpaths)

class Ui_find_id(object):

    def pushButton(self, title, content):
        msgBox = QMessageBox.about(None, title, content)

    def find(self):
        username = self.uname_lineEdit.text()
        phonenum = self.phone_lineEdit.text()
        
        import openpyxl
        wb = openpyxl.load_workbook('회원관리.xlsx')
        ws = wb['회원관리']

        name_exist = False
        phone_exist = False

        # 회원가입 파일에 일치하는 이름이 존재하면 True 반환
        for rowNum in range(2, ws.max_row+1):
            name_value = str(ws.cell(row=rowNum, column=1).value) # 각 행의 이름값
            phone_value = str(ws.cell(row=rowNum, column=4).value) # 각 행의 전화번호

            # 이름과 전화번호 둘 다 일치하면 아이디 반환
            if name_value == username:
                real_rowNum = rowNum
                name_exist = True
            if phone_value == phonenum:
                phone_exist = True

        if name_exist==True and phone_exist==True:
            your_Id = ws.cell(row=real_rowNum, column=2).value
            content = '아이디는 ' + your_Id + '입니다'
            self.pushButton('아이디 찾기', content)
            return
        elif name_exist==True and phone_exist==False:
            self.pushButton('아이디 찾기', '전화번호를 다시 입력하세요.')
            return
        elif name_exist==False and phone_exist==True:
            self.pushButton('아이디 찾기', '이름을 다시 입력하세요.')
            return
        else:
            self.pushButton('아이디 찾기', '일치하는 아이디가 없습니다.')
            return

    def setupUi(self, find_id):
        find_id.setObjectName("find_id")
        find_id.resize(1053, 577)
        self.centralwidget = QtWidgets.QWidget(find_id)
        self.centralwidget.setObjectName("centralwidget")
        self.phone_label = QtWidgets.QLabel(self.centralwidget)
        self.phone_label.setGeometry(QtCore.QRect(170, 320, 171, 91))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.phone_label.setFont(font)
        self.phone_label.setObjectName("phone_label")
        self.uname_label = QtWidgets.QLabel(self.centralwidget)
        self.uname_label.setGeometry(QtCore.QRect(180, 240, 161, 61))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.uname_label.setFont(font)
        self.uname_label.setObjectName("uname_label")
        self.uname_lineEdit = QLineEdit(self.centralwidget)
        self.uname_lineEdit.setGeometry(QtCore.QRect(380, 240, 341, 51))
        self.uname_lineEdit.setObjectName("uname_lineEdit")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(310, 120, 421, 81))
        font = QtGui.QFont()
        font.setPointSize(23)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.phone_lineEdit = QLineEdit(self.centralwidget)
        self.phone_lineEdit.setGeometry(QtCore.QRect(380, 340, 341, 51))
        self.phone_lineEdit.setObjectName("phone_lineEdit")
        self.find_btn = QtWidgets.QPushButton(self.centralwidget)
        self.find_btn.setGeometry(QtCore.QRect(760, 270, 93, 91))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.find_btn.setFont(font)
        self.find_btn.setObjectName("find_btn")
        ###################### 버튼 이벤트 ####################
        self.find_btn.clicked.connect(self.find)
        #####################################################
        find_id.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(find_id)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1053, 26))
        self.menubar.setObjectName("menubar")
        find_id.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(find_id)
        self.statusbar.setObjectName("statusbar")
        find_id.setStatusBar(self.statusbar)

        self.retranslateUi(find_id)
        QtCore.QMetaObject.connectSlotsByName(find_id)

    def retranslateUi(self, find_id):
        _translate = QtCore.QCoreApplication.translate
        find_id.setWindowTitle(_translate("find_id", "MainWindow"))
        self.phone_label.setText(_translate("find_id", "전화번호"))
        self.uname_label.setText(_translate("find_id", "이름"))
        self.label.setText(_translate("find_id", "아이디 찾기"))
        self.find_btn.setText(_translate("find_id", "찾기"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    find_id = QtWidgets.QMainWindow()
    ui = Ui_find_id()
    ui.setupUi(find_id)
    find_id.show()
    sys.exit(app.exec_())

