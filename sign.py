# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'sign.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QMessageBox
'''
from openpyxl import load_workbook
import re
'''
libpaths = QtWidgets.QApplication.libraryPaths() #추가
libpaths.append("C:\\Users\최영주\AppData\Local\Programs\Python\Python37-32\Lib\site-packages\PyQt5\Qt\plugins")
QtWidgets.QApplication.setLibraryPaths(libpaths)
'''
wb = load_workbook('회원관리.xlsx')
ws= wb["회원관리"]
'''
class Ui_sign_up(object):

    def pushButton(self, title, content):
        msgBox = QMessageBox.about(None, title, content)

    def check_id(self, id):
        
        from openpyxl import load_workbook
        import re
        wb = load_workbook('회원관리.xlsx')
        ws= wb["회원관리"]
        
        for rowNum in range(2, ws.max_row+1): #1행은 헤더이므로 제외
            id_value = ws.cell(row=rowNum, column=2).value  #(B(2)열 : 아이디)
            if id_value == id :
                self.pushButton('아이디 중복확인', '이미 존재하는 아이디입니다.')
                return
            else :
                continue
        return False
      
    def validate_id(self):
        id = self.id_edit.text()

        from openpyxl import load_workbook
        import re
        wb = load_workbook('회원관리.xlsx')
        ws= wb["회원관리"]
        
        if len(id) <= 5:
            self.pushButton('아이디 확인', '아이디는 최소 6글자 이상이어야 합니다.')
        elif id.isalnum()!=True :
            self.pushButton('아이디 확인', '아이디는 숫자, 영문자로만 구성되어야 합니다.')
        else:
            if self.check_id(id)==False:
                self.pushButton('아이디 확인', '사용 가능한 아이디입니다.')
                return
            # id = self.validate_id()
        

    def validate_pass(self):
        pwd = self.pass_edit.text()

        from openpyxl import load_workbook
        import re
        wb = load_workbook('회원관리.xlsx')
        ws= wb["회원관리"]
        
        if len(pwd) < 8:
            self.pushButton('비밀번호 확인', '비밀번호는 최소 8글자 이상이어야 합니다.')
        elif re.search('[0-9]',pwd) is None:
            self.pushButton('비밀번호 확인', '비밀번호는 숫자를 최소 하나 이상 포함해야 합니다.')
        elif re.search('[a-zA-Z]',pwd) is None:
            self.pushButton('비밀번호 확인', '비밀번호는 알파벳을 최소 하나 이상 포함해야 합니다.')
        else:
            self.pushButton('비밀번호 확인', '비밀번호가 설정되었습니다.')
    
    def save_info(self, new_list):
        
        from openpyxl import load_workbook
        import re
        wb = load_workbook("회원관리.xlsx")
        ws = wb.active
        
        new_name = self.name_edit.text()
        new_id = self.id_edit.text()
        new_pwd = self.pass_edit.text()
        new_p_number = self.phone_edit.text()
        new_f_used = self.url_edit.text()

        r = str(ws.max_row+1)

        ws['A'+ r] = new_name
        ws['B'+ r] = new_id
        ws['C'+ r] = new_pwd
        ws['D'+ r] = new_p_number
        ws['E'+ r] = new_f_used
        ws['F'+ r] = 0

        self.pushButton('회원가입', '회원가입이 완료되었습니다.')
        wb.save("회원관리.xlsx")

    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(606, 524)
        self.label = QtWidgets.QLabel(Dialog)
        self.label.setGeometry(QtCore.QRect(260, 70, 111, 41))
        font = QtGui.QFont()
        font.setFamily("Agency FB")
        font.setPointSize(20)
        font.setBold(True)
        font.setItalic(False)
        font.setWeight(75)
        font.setStrikeOut(False)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.name_label = QtWidgets.QLabel(Dialog)
        self.name_label.setGeometry(QtCore.QRect(90, 180, 51, 21))
        font = QtGui.QFont()
        font.setPointSize(13)
        self.name_label.setFont(font)
        self.name_label.setObjectName("name_label")
        self.id_label = QtWidgets.QLabel(Dialog)
        self.id_label.setGeometry(QtCore.QRect(90, 220, 61, 21))
        font = QtGui.QFont()
        font.setPointSize(13)
        self.id_label.setFont(font)
        self.id_label.setObjectName("id_label")
        self.pass_label = QtWidgets.QLabel(Dialog)
        self.pass_label.setGeometry(QtCore.QRect(90, 260, 81, 21))
        font = QtGui.QFont()
        font.setPointSize(13)
        self.pass_label.setFont(font)
        self.pass_label.setObjectName("pass_label")
        self.phone_label = QtWidgets.QLabel(Dialog)
        self.phone_label.setGeometry(QtCore.QRect(90, 300, 81, 21))
        font = QtGui.QFont()
        font.setPointSize(13)
        self.phone_label.setFont(font)
        self.phone_label.setObjectName("phone_label")
        self.url_label = QtWidgets.QLabel(Dialog)
        self.url_label.setGeometry(QtCore.QRect(90, 350, 181, 21))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.url_label.setFont(font)
        self.url_label.setObjectName("url_label")
        self.name_edit = QtWidgets.QLineEdit(Dialog)
        self.name_edit.setGeometry(QtCore.QRect(260, 180, 171, 21))
        self.name_edit.setObjectName("name_edit")
        self.id_edit = QtWidgets.QLineEdit(Dialog)
        self.id_edit.setGeometry(QtCore.QRect(260, 220, 171, 21))
        self.id_edit.setObjectName("id_edit")
        self.pass_edit = QtWidgets.QLineEdit(Dialog)
        self.pass_edit.setGeometry(QtCore.QRect(260, 260, 171, 21))
        self.pass_edit.setObjectName("pass_edit")
        self.phone_edit = QtWidgets.QLineEdit(Dialog)
        self.phone_edit.setGeometry(QtCore.QRect(260, 300, 171, 21))
        self.phone_edit.setObjectName("phone_edit")
        self.url_edit = QtWidgets.QLineEdit(Dialog)
        self.url_edit.setGeometry(QtCore.QRect(260, 350, 171, 21))
        self.url_edit.setObjectName("url_edit")
        
        self.checkId_btn = QtWidgets.QPushButton(Dialog)
        self.checkId_btn.setGeometry(QtCore.QRect(470, 220, 91, 21))        
        self.checkId_btn.setObjectName("checkId_btn")
        # 아이디 확인 버튼의 이벤트
        self.checkId_btn.clicked.connect(self.validate_id)
        
        self.checkPass_btn = QtWidgets.QPushButton(Dialog)
        self.checkPass_btn.setGeometry(QtCore.QRect(470, 260, 91, 21))       
        self.checkPass_btn.setObjectName("checkPass_btn")
        # 비밀번호 확인 버튼의 이벤트
        self.checkPass_btn.clicked.connect(self.validate_pass)

        self.sign_btn = QtWidgets.QPushButton(Dialog)
        self.sign_btn.setGeometry(QtCore.QRect(260, 440, 91, 31))
        self.sign_btn.setObjectName("sign_btn")
        # 가입하기 버튼의 이벤트
        self.sign_btn.clicked.connect(self.save_info)        
        
        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Dialog"))
        self.label.setText(_translate("Dialog", "회원가입"))
        self.name_label.setText(_translate("Dialog", "이름"))
        self.id_label.setText(_translate("Dialog", "아이디"))
        self.pass_label.setText(_translate("Dialog", "비밀번호"))
        self.phone_label.setText(_translate("Dialog", "전화번호"))
        self.url_label.setText(_translate("Dialog", "자주 사용하는 종목"))
        self.sign_btn.setText(_translate("Dialog", "가입하기"))
        self.checkId_btn.setText(_translate("Dialog", "아이디 확인"))
        self.checkPass_btn.setText(_translate("Dialog", "비밀번호 확인"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_sign_up()
    ui.setupUi(Dialog)
    Dialog.show()
    sys.exit(app.exec_())

