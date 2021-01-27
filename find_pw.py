# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'find_pw.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *

libpaths = QApplication.libraryPaths() 
libpaths.append("C:\\Users\김민휘\AppData\Local\Programs\Python\Python36-32\Lib\site-packages\PyQt5\Qt\plugins") 
QApplication.setLibraryPaths(libpaths)

class Ui_find_pw(object):

    def pushButton(self, title, content):
        msgBox = QMessageBox.about(None, title, content)

    def find(self):
        username = self.uname_lineEdit.text()
        userid = self.id_lineEdit.text()

        import openpyxl
        wb = openpyxl.load_workbook('회원관리.xlsx')
        ws = wb['회원관리']

        id_exist = False
        name_exist = False

         # 회원가입 파일에 일치하는 이름이 존재하면 True 반환
        for rowNum in range(2, ws.max_row+1):
            id_value = str(ws.cell(row=rowNum, column=2).value) # 각 행의 이름값
            name_value = str(ws.cell(row=rowNum, column=1).value) # 각 행의 전화번호

            # 이름과 전화번호 둘 다 일치하면 아이디 반환
            if id_value == userid:
                real_rowNum = rowNum
                id_exist = True
            if name_value == username:
                name_exist = True

        if id_exist==True and name_exist==True:
            your_password = ws.cell(row=real_rowNum, column=3).value
            content = '비밀번호는 ' + your_password + '입니다'
            self.pushButton('비밀번호 찾기', content)
            return
        elif id_exist==True and name_exist==False:
            self.pushButton('비밀번호 찾기',"이름을 다시 입력하세요.")
            return
        elif id_exist==False and name_exist==True:
            self.pushButton('비밀번호 찾기',"아이디를 다시 입력하세요.")
            return
        else:
            self.pushButton('비밀번호 찾기',"일치하는 패스워드가 없습니다.")
            return

    def setupUi(self, find_pw):
        find_pw.setObjectName("find_pw")
        find_pw.resize(996, 673)
        self.centralwidget = QtWidgets.QWidget(find_pw)
        self.centralwidget.setObjectName("centralwidget")
        self.id_label = QtWidgets.QLabel(self.centralwidget)
        self.id_label.setGeometry(QtCore.QRect(160, 290, 161, 61))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.id_label.setFont(font)
        self.id_label.setObjectName("id_label")
        self.uname_label = QtWidgets.QLabel(self.centralwidget)
        self.uname_label.setGeometry(QtCore.QRect(150, 370, 171, 91))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.uname_label.setFont(font)
        self.uname_label.setObjectName("uname_label")
        self.id_lineEdit = QLineEdit(self.centralwidget)
        self.id_lineEdit.setGeometry(QtCore.QRect(360, 290, 341, 51))
        self.id_lineEdit.setObjectName("id_lineEdit")
        self.find_btn = QtWidgets.QPushButton(self.centralwidget)
        self.find_btn.setGeometry(QtCore.QRect(740, 320, 93, 91))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.find_btn.setFont(font)
        self.find_btn.setObjectName("find_btn")
        ###################### 버튼 이벤트 ####################
        self.find_btn.clicked.connect(self.find)
        #####################################################
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(290, 170, 421, 81))
        font = QtGui.QFont()
        font.setPointSize(23)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.uname_lineEdit = QLineEdit(self.centralwidget)
        self.uname_lineEdit.setGeometry(QtCore.QRect(360, 390, 341, 51))
        self.uname_lineEdit.setObjectName("uname_lineEdit")
        find_pw.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(find_pw)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 996, 26))
        self.menubar.setObjectName("menubar")
        find_pw.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(find_pw)
        self.statusbar.setObjectName("statusbar")
        find_pw.setStatusBar(self.statusbar)

        self.retranslateUi(find_pw)
        QtCore.QMetaObject.connectSlotsByName(find_pw)

    def retranslateUi(self, find_pw):
        _translate = QtCore.QCoreApplication.translate
        find_pw.setWindowTitle(_translate("find_pw", "MainWindow"))
        self.id_label.setText(_translate("find_pw", "아이디"))
        self.uname_label.setText(_translate("find_pw", "이름"))
        self.find_btn.setText(_translate("find_pw", "찾기"))
        self.label.setText(_translate("find_pw", "비밀번호 찾기"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    find_pw = QtWidgets.QMainWindow()
    ui = Ui_find_pw()
    ui.setupUi(find_pw)
    find_pw.show()
    sys.exit(app.exec_())

