# 시간충전 pyqt gui 구현 (이주연)
# 1. user_id 불러오는 부분(login시 넘겨줘야 함.)
# 2. dict_cost 넘겨주는 부분 main함수에 통합할때 고려해야함.

# 주문단계와 연결
from order_product import *

# 회원관리 엑셀 파일 불러오기

import openpyxl
import re 
wb = openpyxl.load_workbook('회원관리.xlsx')
ws = wb['회원관리']

# 결제로그 엑셀 파일 불러오기

wb2 = openpyxl.load_workbook('결제로그.xlsx')
ws2 = wb2['결제로그']

# 결제로그 파일에서 읽어온 현재 사용자의 아이디 저장
user_id = ws2.cell(row = ws2.max_row, column=2).value 
time, cost = 0, 0 # 전역변수 

import sys #
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import pyqtSlot #
from PyQt5.QtGui import QIcon #
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QMessageBox #

class Ui_ChargeTime(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(600, 550)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(50, 50, 170, 60))
        self.pushButton.setObjectName("pushButton") # [1]. 1시간 충전
        self.pushButton.clicked.connect(self.identify_time)
        
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(50, 150, 170, 60))    
        self.pushButton_2.setObjectName("pushButton_2") # [2]. 3시간 충전
        self.pushButton_2.clicked.connect(self.identify_time2)
        
        self.pushButton_3 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_3.setGeometry(QtCore.QRect(50, 250, 170, 60))
        self.pushButton_3.setObjectName("pushButton_3") # [3]. 7시간 충전
        self.pushButton_3.clicked.connect(self.identify_time3)
        
        self.pushButton_4 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_4.setGeometry(QtCore.QRect(20, 460, 90, 40))
        self.pushButton_4.setObjectName("pushButton_4") # 이전 버튼
        
        self.pushButton_5 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_5.setGeometry(QtCore.QRect(470, 460, 90, 40))
        self.pushButton_5.setObjectName("pushButton_5") # 다음 버튼
        self.pushButton_5.clicked.connect(self.dict_cost)#dict_cost
        #print(a)
        self.pushButton_5.clicked.connect(self.next_button)
        
        self.pushButton_6 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_6.setGeometry(QtCore.QRect(470, 410, 90, 40))
        self.pushButton_6.setObjectName("pushButton_6") # 삭제 버튼 
        self.pushButton_6.clicked.connect(self.delete_item)
        '''
        self.textEdit = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit.setGeometry(QtCore.QRect(260, 400, 70, 20))
        self.textEdit.setObjectName("textEdit") #안내문 출력
        #self.textEdit.setText(">>결제를 원하시면,하단의 '다음'을 눌러주세요.")
        '''
        self.listWidget = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget.setGeometry(QtCore.QRect(260, 50, 280, 260))
        self.listWidget.setObjectName("listWidget") # 리스트에 시간추가

        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(260, 340, 200, 30))
        self.label.setObjectName("label") # 충전 전 잔여시간 표시 
        charge_before_time = self.charge_before_time()
        self.label.setText('<충전 전 잔여시간>: %d' %(charge_before_time))
        
        self.label2 = QtWidgets.QLabel(self.centralwidget)
        self.label2.setGeometry(QtCore.QRect(260, 370, 200, 30))
        self.label2.setObjectName("label2")# 충전 후 잔여시간 표시

        self.label3 = QtWidgets.QLabel(self.centralwidget)
        self.label3.setGeometry(QtCore.QRect(260, 400, 200, 30))
        self.label3.setObjectName("label3")# 총 금액 표시
        
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 430, 18))
        
        
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "시간충전"))
        self.pushButton.setText(_translate("MainWindow", "[1] 1시간:  1000원"))
        self.pushButton_2.setText(_translate("MainWindow", "[2] 3시간: 2000원"))
        self.pushButton_3.setText(_translate("MainWindow", "[3] 7시간: 5000원"))
        self.pushButton_4.setText(_translate("MainWindow", "<< 이전"))
        self.pushButton_5.setText(_translate("MainWindow", "다음 >>"))
        self.pushButton_6.setText(_translate("MainWindow", "삭제"))
        #self.label.setText(_translate("MainWindow", "TextLabel"))
        

    def charge_before_time(self): # 충전 전 잔여시간 구하기
        global time

        import openpyxl
        import re 
        wb = openpyxl.load_workbook('회원관리.xlsx')
        ws = wb['회원관리']
        
        for rowNum in range(2, ws.max_row+1) :   
            id_value = str(ws.cell(row=rowNum, column=2).value)
            if id_value == user_id :
                break
        charge_before_time = ws.cell(row=rowNum, column=6).value
        time = charge_before_time
        return charge_before_time
    
    def current_time_cost(self):
        self.label2.setText('<충전 후 잔여시간>: %d' %(time))
        self.label3.setText('**총 금액: %d' %(cost))
        
    def dict_cost(self):
        import openpyxl
        import re 
        wb2 = openpyxl.load_workbook('결제로그.xlsx')
        ws2 = wb2['결제로그']
        
        print("충전 후 잔여시간은 [%d시간]입니다.\n" %(time))
        
        ws2.cell(row = ws2.max_row, column = 3).value = time
        ws2.cell(row = ws2.max_row, column = 5).value = cost
        wb2.save('결제로그.xlsx')
        
    def next_button(self):
        
        msgBox = QMessageBox.about(None,'안내문','상품주문으로 넘어갑니다.')
    
        self.window = QtWidgets.QDialog()
        self.ui = Ui_OrderProduct()
        self.ui.setupUi(self.window)
        self.window.show()
        
        #msgBox.exec_()
        
    '''    
    def print_content(self,time,cost,num):
        self.textEdit.clear()
        self.textEdit.append("[%d]번이 선택되었습니다.\n" %num)
        self.textEdit.append("충전시간: %d시간\n" %time)
        self.textEdit.append("금액: %d원\n\n" %cost)
        self.textEdit.append(">>[%d]번 선택을 원하시면, 하단의 '다음'을 눌러주세요." %num)
    '''
    def add_time(self,t,c):
        global time
        global cost
        self.listWidget.addItem('충전시간: %d시간 / 금액: %d원' %(t,c))
        time += t
        cost += c
        self.current_time_cost()

    def delete_item(self):
        global time
        global cost
        listItems = self.listWidget.selectedItems()
        if not listItems:
            msgBox = QMessageBox.about(None,'오류','삭제할 시간을 선택해주세요.')
            return
        s = self.listWidget.currentItem().text()
        s = re.split('\W+', s)       
        time -= int(s[1][0])
        cost -= int(s[3][:-1])
        
        for item in listItems:
            self.listWidget.takeItem(self.listWidget.row(item))
            
        self.current_time_cost()
            
    def identify_time(self):
        time = 1
        cost = 1000
        print('1번 시간충전')
        self.add_time(time,cost)

    def identify_time2(self):
        time = 3
        cost = 2000
        print('2번 시간충전')
        self.add_time(time,cost)
        
    def identify_time3(self):
        time = 7
        cost = 5000
        print('3번 시간충전')
        self.add_time(time,cost)
        
if __name__ == "__main__":
    import sys
    libpaths = QtWidgets.QApplication.libraryPaths() #추가
    libpaths.append("C:\\Users\사용자\AppData\Local\Programs\Python\Python37-32\Lib\site-packages\PyQt5\Qt\plugins")
    QtWidgets.QApplication.setLibraryPaths(libpaths)
    
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_ChargeTime()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

