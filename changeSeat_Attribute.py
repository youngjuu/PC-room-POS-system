# 좌석속성설기능
# 최영주

from openpyxl import load_workbook

wb = load_workbook("좌석관리.xlsx")
ws = wb.active

# 좌석 속성 바꾸는 함수
def changeAttribute():
    while True:
        seat_number = int(input("속성 변경을 원하는 좌석번호를 입력하세요(1~30)\n좌석 속성 변경을 종료하려면 0을 입력하세요. "))
        if (seat_number>30):
            print("1~30사이의 좌석번호를 입력해주세요.\n")
            continue
        if seat_number == 0:
            print("좌석 속성 변경을 종료합니다.")
            break
        
        want_attribute = input("원하는 속성을 선택하세요.\n일반석,커플석,고사양게임석,티켓팅석 중 하나를 입력하세요 ")
        print()
        ws.cell(row=seat_number+1, column=2).value = want_attribute

    wb.save("좌석관리.xlsx")

#changeAttribute()
