# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'login.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from find_id import *
from find_pw import *
from admin_window import *
from next_login import *


libpaths = QtWidgets.QApplication.libraryPaths() #추가
libpaths.append("C:\\Users\사용자\AppData\Local\Programs\Python\Python37-32\Lib\site-packages\PyQt5\Qt\plugins")
QtWidgets.QApplication.setLibraryPaths(libpaths)
'''
from openpyxl import load_workbook
wb = load_workbook("회원관리.xlsx")
ws = wb.active
wb1 = load_workbook("결제로그.xlsx")
ws1 = wb1.active
'''

class Ui_Dialog(object):

    def pushButton(self, title, content):
        msgBox = QMessageBox.about(None, title, content)
    
    def login(self):
        username = self.uname_lineEdit.text()
        password = self.pass_lineEdit.text()
        #추가 
        from openpyxl import load_workbook
        wb = load_workbook("회원관리.xlsx")
        ws = wb.active
        wb1 = load_workbook("결제로그.xlsx")
        ws1 = wb1.active

        # 관리자 로그인 성공
        if (username == "admin") and (password == "0000"):
            self.pushButton('로그인 성공', '관리자 로그인에 성공했습니다.')

            r = str(ws1.max_row+1)
            ws1['B' + r] = username
            wb1.save("결제로그.xlsx")
            
            self.window = QtWidgets.QMainWindow()
            self.ui = Ui_adminWindow()
            self.ui.setupUi(self.window)
            self.window.show()
            return


        # 사용자 로그인 성공
        for rowNum in range(2, ws.max_row+1) :
            if (username == str(ws.cell(row=rowNum, column=2).value)) and (password == str(ws.cell(row=rowNum, column=3).value)) :
                often_use = ws.cell(row=rowNum, column=5).value
                rest = ws.cell(row=rowNum, column=6).value
                content = '로그인에 성공했습니다.\n잔여시간은 %d시간 입니다.\n자주 사용하는 %s URL 주소를 PC에 전송해 놓겠습니다.\n' %(rest, often_use)
                self.pushButton('로그인 성공', content)
                
                r = str(ws1.max_row+1)
                ws1['B' + r] = username
                wb1.save("결제로그.xlsx")
                return

        # 로그인 실패
        self.pushButton('로그인 실패', '로그인에 실패했습니다.')
        return
               
    def find_id(self) :
        self.window = QtWidgets.QMainWindow()
        self.ui = Ui_find_id()
        self.ui.setupUi(self.window)
        self.window.show()

    def find_pw(self) :
        self.window = QtWidgets.QMainWindow()
        self.ui = Ui_find_pw()
        self.ui.setupUi(self.window)
        self.window.show()

    def next(self) :

        username = self.uname_lineEdit.text()
        password = self.pass_lineEdit.text()

        #추가 
        from openpyxl import load_workbook
        wb = load_workbook("회원관리.xlsx")
        ws = wb.active
        wb1 = load_workbook("결제로그.xlsx")
        ws1 = wb1.active

        last_user = ws1.cell(row = ws1.max_row, column=2).value

        for rowNum in range(2, ws.max_row+1):
            if last_user == str(ws.cell(row=rowNum, column=2).value):
                                break
                               
        last_password = ws.cell(row=rowNum, column=3).value
                                       
        if username == last_user and password == last_password:
            if ws1.cell(row = ws1.max_row, column=2).value  == 'admin' :
                self.ui = Ui_adminWindow()
                self.ui.setupUi(self.window)
                self.window.show()

            else :
                self.window = QtWidgets.QMainWindow()
                self.ui = Ui_MainWindow()
                self.ui.setupUi(self.window)
                self.window.show()
                
        # 로그인 실패 시 다음 단계로 넘어가지 못함
        else:
            self.pushButton('안내문', '먼저 로그인을 하십시오.')
            return
        
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(1094, 794)
        self.uname_label = QtWidgets.QLabel(Dialog)
        self.uname_label.setGeometry(QtCore.QRect(270, 280, 161, 61))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.uname_label.setFont(font)
        self.uname_label.setObjectName("uname_label")
        self.pass_label = QtWidgets.QLabel(Dialog)
        self.pass_label.setGeometry(QtCore.QRect(270, 360, 171, 91))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.pass_label.setFont(font)
        self.pass_label.setObjectName("pass_label")
        self.uname_lineEdit = QLineEdit(Dialog)
        self.uname_lineEdit.setGeometry(QtCore.QRect(420, 280, 341, 51))
        self.uname_lineEdit.setObjectName("uname_lineEdit")
        self.pass_lineEdit = QLineEdit(Dialog)
        self.pass_lineEdit.setGeometry(QtCore.QRect(420, 370, 341, 51))
        self.pass_lineEdit.setObjectName("pass_lineEdit")
        self.login_btn = QtWidgets.QPushButton(Dialog)
        self.login_btn.setGeometry(QtCore.QRect(790, 310, 93, 91))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.login_btn.setFont(font)
        self.login_btn.setObjectName("login_btn")
        ###################### 버튼 이벤트 ####################
        self.login_btn.clicked.connect(self.login)
        #####################################################
        self.find_id_btn = QtWidgets.QPushButton(Dialog)
        self.find_id_btn.setGeometry(QtCore.QRect(470, 470, 93, 28))
        self.find_id_btn.setObjectName("find_id_btn")
        ###################### 버튼 이벤트 ####################
        self.find_id_btn.clicked.connect(self.find_id)
        #####################################################
        self.find_pw_btn = QtWidgets.QPushButton(Dialog)
        self.find_pw_btn.setGeometry(QtCore.QRect(580, 470, 111, 28))
        self.find_pw_btn.setObjectName("find_pw_btn")
         ###################### 버튼 이벤트 ####################
        self.find_pw_btn.clicked.connect(self.find_pw)
        #####################################################
        self.next_btn = QtWidgets.QPushButton(Dialog)
        self.next_btn.setGeometry(QtCore.QRect(790, 470, 90, 28))
        self.next_btn.setObjectName("next_btn")
        ###################### 버튼 이벤트 ###################
        self.next_btn.clicked.connect(self.next)
        ########################################################
        
        self.label = QtWidgets.QLabel(Dialog)
        self.label.setGeometry(QtCore.QRect(510, 160, 421, 81))
        font = QtGui.QFont()
        font.setPointSize(23)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "로그인"))
        self.uname_label.setText(_translate("Dialog", "아이디"))
        self.pass_label.setText(_translate("Dialog", "비밀번호"))
        self.login_btn.setText(_translate("Dialog", "로그인"))
        self.find_id_btn.setText(_translate("Dialog", "아이디 찾기"))
        self.find_pw_btn.setText(_translate("Dialog", "비밀번호 찾기"))
        self.next_btn.setText(_translate("Dialog", "다음 >>"))
        self.label.setText(_translate("Dialog", "로그인"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.show()
    sys.exit(app.exec_())

