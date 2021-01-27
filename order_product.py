# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'order_product.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
import sys
'''
from openpyxl import load_workbook
wb = load_workbook("상품관리.xlsx")
ws = wb.active
wl = load_workbook('결제로그.xlsx')
wls = wl.active
'''
from lastOrder import * # 결제단계와 연결
libpaths = QtWidgets.QApplication.libraryPaths() #추가
libpaths.append("C:\\Users\사용자\AppData\Local\Programs\Python\Python37-32\Lib\site-packages\PyQt5\Qt\plugins")
QtWidgets.QApplication.setLibraryPaths(libpaths)

class Ui_OrderProduct(object):
#------------------------------------------------------------------------------------------------------------------
#이벤트 헨들러
    #슬롯 함수
    def text_show_price(self,dict_menucost):
        product=self.product_list.currentText()
        self.show_price.setText('%d원' %(dict_menucost[product][0]))

    def text_show_leftover(self,dict_menucost):
        product=self.product_list.currentText()
        self.show_leftover.setText('%d개' %(dict_menucost[product][1]))
        self.order_amount.setMaximum(dict_menucost[product][1]) #최대 수량 = 재고량
        #count=self.order_amount.value()
        
    def text_add_order(self,dict_menucost):
        index = self.product_list.currentIndex()
        menu=list(dict_menucost.keys())
        count=self.order_amount.value()
        self.total_order_check.addItem('%s : %d개' %(menu[index], count))
        product=self.product_list.currentText()
        cost=self.show_total_amount.text()[:-1]
        cost=int(cost)
        cost+= (dict_menucost[product][0]*count)        
        self.show_total_amount.setText('%d원' %cost)
        dict_menucost[product][1] = dict_menucost[product][1]-count
        self.order_amount.setMaximum(dict_menucost[product][1])#최대 수량 = 재고량
        self.show_leftover.setText('%d개' %(dict_menucost[product][1]))
        
    def text_delete_order(self,dict_menucost):
        product=self.product_list.currentText()
        listItems=self.total_order_check.selectedItems()
        if not listItems:
            msgBox = QtWidgets.QMessageBox.about(None,"오류","삭제할 상품을 선택해주세요")
            #print('상품을 선택해주세요')
            return
        s = self.total_order_check.currentItem().text().split(' : ')
        cost=self.show_total_amount.text()[:-1]
        cost=int(cost)
        cost-= (dict_menucost[s[0]][0])*(int(s[1][:-1]))        
        self.show_total_amount.setText('%d원' %cost)
        dict_menucost[product][1]=dict_menucost[product][1]+int(s[1][:-1])
        self.order_amount.setMaximum(dict_menucost[product][1]) #최대 수량 = 재고량
        self.show_leftover.setText('%d개' %(dict_menucost[product][1]))
        for item in listItems:
            self.total_order_check.takeItem(self.total_order_check.row(item))
            
    def text_now_sale(self,sale_menu):
        self.now_sale.clear()
        self.now_sale.addItems(sale_menu)
#------------------------------------------------------------------------------------------------------------------        
#엑셀(결제로그)로 보낼 내용
    def orderButtonClicked(self,dict_menucost):
        from openpyxl import load_workbook
        wb = load_workbook("상품관리.xlsx")
        ws = wb.active
        wl = load_workbook('결제로그.xlsx')
        wls = wl.active
        
        product_bought=[]
        for i in dict_menucost:
            product_bought.append(dict_menucost[i][1])
        product_bought="/".join(map(str,product_bought))
        #print(product_bought)
        r = str(wls.max_row)

        wls['D'+ r] = product_bought
        wls['F'+ r] = int(self.show_total_amount.text()[:-1])
        wl.save('결제로그.xlsx')
        msgBox = QtWidgets.QMessageBox.about(None,"주문완료","주문이 완료되었습니다.")

        self.window = QtWidgets.QMainWindow() # 결제단계와 연결
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self.window)
        self.window.show()
        
#------------------------------------------------------------------------------------------------------------------        
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(521, 504)
        self.product_list = QtWidgets.QComboBox(Form)
        self.product_list.setGeometry(QtCore.QRect(130, 90, 141, 21))
        self.product_list.setObjectName("product_list")
        self.lbl_product = QtWidgets.QLabel(Form)
        self.lbl_product.setGeometry(QtCore.QRect(40, 90, 71, 21))
        self.lbl_product.setAlignment(QtCore.Qt.AlignCenter)
        self.lbl_product.setObjectName("lbl_product")
        self.lbl_price = QtWidgets.QLabel(Form)
        self.lbl_price.setGeometry(QtCore.QRect(40, 140, 71, 21))
        self.lbl_price.setAlignment(QtCore.Qt.AlignCenter)
        self.lbl_price.setObjectName("lbl_price")
        self.lbl_leftover = QtWidgets.QLabel(Form)
        self.lbl_leftover.setGeometry(QtCore.QRect(40, 190, 71, 21))
        self.lbl_leftover.setAlignment(QtCore.Qt.AlignCenter)
        self.lbl_leftover.setObjectName("lbl_leftover")
        self.lbl_order_amount = QtWidgets.QLabel(Form)
        self.lbl_order_amount.setGeometry(QtCore.QRect(40, 240, 101, 21))
        self.lbl_order_amount.setAlignment(QtCore.Qt.AlignCenter)
        self.lbl_order_amount.setObjectName("lbl_order_amount")
        self.order_amount = QtWidgets.QSpinBox(Form)
        self.order_amount.setGeometry(QtCore.QRect(150, 240, 42, 22))
        self.order_amount.setObjectName("order_amount")
        self.bttn_add_product = QtWidgets.QPushButton(Form)
        self.bttn_add_product.setGeometry(QtCore.QRect(230, 240, 111, 31))
        self.bttn_add_product.setObjectName("bttn_add_product")
        self.check_discount = QtWidgets.QCheckBox(Form)
        self.check_discount.setGeometry(QtCore.QRect(310, 80, 141, 41))
        self.check_discount.setObjectName("check_discount")
        self.lbl_total_price = QtWidgets.QLabel(Form)
        self.lbl_total_price.setGeometry(QtCore.QRect(50, 400, 101, 21))
        self.lbl_total_price.setAlignment(QtCore.Qt.AlignCenter)
        self.lbl_total_price.setObjectName("lbl_total_price")
        self.bttn_order = QtWidgets.QPushButton(Form)
        self.bttn_order.setGeometry(QtCore.QRect(280, 396, 111, 30))
        self.bttn_order.setObjectName("bttn_order")
        self.show_price = QtWidgets.QLabel(Form)
        self.show_price.setGeometry(QtCore.QRect(130, 140, 113, 20))
        self.show_price.setObjectName("show_price")
        self.show_leftover = QtWidgets.QLabel(Form)
        self.show_leftover.setGeometry(QtCore.QRect(130, 190, 113, 20))
        self.show_leftover.setObjectName("show_leftover")
        self.bttn_delete_product = QtWidgets.QPushButton(Form)
        self.bttn_delete_product.setGeometry(QtCore.QRect(370, 240, 111, 31))
        self.bttn_delete_product.setObjectName("bttn_delete_product")
        self.total_order_check = QtWidgets.QListWidget(Form)
        self.total_order_check.setGeometry(QtCore.QRect(130, 290, 251, 81))
        self.total_order_check.setObjectName("total_order_check")
        self.lbl_total_order_check = QtWidgets.QLabel(Form)
        self.lbl_total_order_check.setGeometry(QtCore.QRect(40, 320, 101, 21))
        self.lbl_total_order_check.setAlignment(QtCore.Qt.AlignCenter)
        self.lbl_total_order_check.setObjectName("lbl_total_order_check")
        self.show_total_amount = QtWidgets.QLabel(Form)
        self.show_total_amount.setGeometry(QtCore.QRect(150, 400, 111, 21))
        self.show_total_amount.setObjectName("show_total_amount")
        self.now_sale = QtWidgets.QListWidget(Form)
        self.now_sale.setGeometry(QtCore.QRect(309, 113, 111, 81))
        self.now_sale.setObjectName("now_sale")

        self.retranslateUi(Form)
#------------------------------------------------------------------------------------------------------------------
    #데이터 설정
    #콤보 상자에 상품 설정
        dict_menucost = {}
        menu=[]
        sale_menu=[]
        cost= 0
        
        from openpyxl import load_workbook
        wb = load_workbook("상품관리.xlsx")
        ws = wb.active
        wl = load_workbook('결제로그.xlsx')
        wls = wl.active
        for rowNum in range(2, ws.max_row+1) :

            # 현재 할인상품이 아닌 경우
            if ws.cell(row=rowNum, column=4).value == 'NO' :
                dict_menucost[ws.cell(row=rowNum, column=1).value] = [ws.cell(row=rowNum, column=3).value,ws.cell(row=rowNum, column=7).value]
                menu.append(ws.cell(row=rowNum,column=1).value)
            # 현재 할인상품인 경우  
            elif ws.cell(row=rowNum, column=4).value == 'YES' :
                dict_menucost[ws.cell(row=rowNum, column=1).value] = [ws.cell(row=rowNum, column=5).value,ws.cell(row=rowNum, column=7).value]
                menu.append(ws.cell(row=rowNum,column=1).value)
                sale_menu.append(ws.cell(row=rowNum,column=1).value)
            #dict_menuamount[ws.cell(row=rowNum, column=1).value]=ws.cell(row=rowNum, column=7).value
        #print(dict_menucost)
        #print(menu)

        self.product_list.addItems(menu)
        
    #디폴트 가격 표시
        product=self.product_list.currentText()
        self.show_price.setText('%d원' %(dict_menucost[product][0]))

    #디폴트 재고 표시
        row = self.product_list.currentIndex()+2
        self.show_leftover.setText('%d개' %(dict_menucost[product][1]))
        
    #디폴트 주문 수량 설정
        self.order_amount.setMaximum(dict_menucost[product][1]) #최대 수량 = 재고량
        count=self.order_amount.value()
        
    #주문 내역 확인
        #self.total_order_check.addItem('%s : %d개' %(menu[row-2], count))

    #총 합계 금액 설정       
        self.show_total_amount.setText('%d원' %cost)
        
#이벤트 헨들러
    #커넥션
        self.product_list.activated['int'].connect(lambda: self.text_show_price(dict_menucost))             
        self.product_list.activated['int'].connect(lambda: self.text_show_leftover(dict_menucost))
        self.bttn_add_product.clicked.connect(lambda: self.text_add_order(dict_menucost))
        self.bttn_delete_product.clicked.connect(lambda: self.text_delete_order(dict_menucost))
        self.check_discount.clicked['bool'].connect(lambda: self.text_now_sale(sale_menu))
        self.bttn_order.clicked.connect(lambda:self.orderButtonClicked(dict_menucost))
        QtCore.QMetaObject.connectSlotsByName(Form)
    
    
        
        
#------------------------------------------------------------------------------------------------------------------
    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "상품주문"))
        self.lbl_product.setText(_translate("Form", "상품"))
        self.lbl_price.setText(_translate("Form", "가격"))
        self.lbl_leftover.setText(_translate("Form", "재고"))
        self.lbl_order_amount.setText(_translate("Form", "주문 수량"))
        self.bttn_add_product.setText(_translate("Form", "상품 추가"))
        self.check_discount.setText(_translate("Form", "현재 할인 상품"))
        self.lbl_total_price.setText(_translate("Form", "상품 총 합계"))
        self.bttn_order.setText(_translate("Form", "주문하기")) #결제단계와 연결
        self.bttn_delete_product.setText(_translate("Form", "상품 삭제"))
        self.lbl_total_order_check.setText(_translate("Form", "주문 내역"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Form = QtWidgets.QWidget()
    ui = Ui_OrderProduct()
    ui.setupUi(Form)
    Form.show()
    sys.exit(app.exec_())

