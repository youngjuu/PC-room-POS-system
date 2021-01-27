#N. 좌석 선택
#김세영
from openpyxl import load_workbook
wb = load_workbook('좌석관리.xlsx')
ws = wb.active

def select_seat() :
    while True :
        selected_seat = int(input('좌석번호를 선택해 주십시오.(1~30) '))

        if(ws.cell(row= selected_seat + 1, column = 3).value == True):
            print('이미 사용중인 좌석입니다. 다시 선택해주십시오.')
            continue
        else:
            print('%d번 좌석이 선택되었습니다.' %selected_seat)
            return selected_seat
        
def save_seat(user_seat):
    ws.cell(row= user_seat + 1, column = 3).value = True
    wb.save('좌석관리.xlsx')

