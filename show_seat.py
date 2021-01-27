#I 인원수 배분 알고리즘
#김세영

from openpyxl import load_workbook
wb = load_workbook('좌석관리.xlsx')
ws = wb.active

#'좌석관리' 엑셀 파일에서 데이터 읽어오기(str으로 변환 후 seats 리스트에 저장) 
def read_seatdata_fromxlsx():
    #자리 정보(T or F) 담을 리스트 생성
    temp_seats = []
    seats = [] 
    string = ''
    count = 0
    
    for rowNum in range(2,ws.max_row+1):
        seat_data = int(ws.cell(row= rowNum, column = 3).value)
        temp_seats.append(seat_data)
    
    for i in range(len(temp_seats)):
        if( count == 6):
            seats.append(string)
            string = str(temp_seats[i])
            count = 1
        else:
            string += str(temp_seats[i])
            count += 1
    seats.append(string)
    return seats

def check_six (seats) :
    n_ppl = int(input('인원 수를 입력해주세요 '))
    #print('<현재 좌석>')
    #print(seats)
    print()
    
    #6명 이하
    if n_ppl <= 6:
        if search_empty_seats('0',seats,n_ppl) == True :
            print('<추천 좌석>')
            new_seats = show_seats(seats,[n_ppl],n_ppl)
            return new_seats
        else:
            while True :
                continue_flag = False
                ppl = input('[ㅁ/ㅁ/ㅁ] 형태로 작성해 주세요. (최대 6명까지 ㅁ에 입력가능)EX.8명일시에 4/4 혹은 2/2/4 ')
                ppl = ppl.split('/')
                ppl = [int(i) for i in ppl]
              
                #(1) 나눈 인원 수가 총 인원수와 일치하지 않는 경우
                if check_total_num(ppl,n_ppl) == False:
                    print('총 인원수와 일치하도록 다시 입력해주십시오.')
                    print()
                    continue
                
                #(2) 빈 좌석이 없는 경우 
                for i in range(len(ppl)):
                    if search_empty_seats('0',seats,ppl[i]) == False :
                        print('좌석을 찾을 수 없습니다.좌석을 다시 나눠주세요')
                        #print('<현재 좌석>')
                        #print(seats)
                        print()
                        continue_flag = True
                if(continue_flag == True):
                    continue
                
                #(3) 좌석이 존재하는 경우
                print('<추천 좌석>')
                new_seats = show_seats(seats,ppl,n_ppl)
                return new_seats
    #6명 초과        
    elif n_ppl>6 and n_ppl<=30:
        while True :
                continue_flag = False
                ppl = input('[ㅁ/ㅁ/ㅁ] 형태로 작성해 주세요. (최대 6명까지 ㅁ에 입력가능)EX.8명일시에 4/4 혹은 2/2/4 ')
                ppl = ppl.split('/')
                ppl = [int(i) for i in ppl]

                #(1) 나눈 인원 수가 총 인원수와 일치하지 않는 경우
                if check_total_num(ppl,n_ppl) == False:
                    print('총 인원수와 일치하도록 다시 입력해주십시오.')
                    print()
                    continue

                #(2) 빈 좌석이 없는 경우 
                for i in range(len(ppl)):
                    if search_empty_seats('0',seats,ppl[i]) == False :
                        print('좌석을 찾을 수 없습니다.좌석을 다시 나눠주세요.')
                        #print('<현재 좌석>')
                        #print(seats)
                        print()
                        continue_flag = True
                        
                if(continue_flag == True):
                    continue
                
                #(3) 좌석이 존재하는 경우
                print('<추천 좌석>')
                new_seats = show_seats(seats,ppl,n_ppl)
                print(new_seats)
                return new_seats

#리스트 안의 값을 string으로 변환
def list_value_into_string(list):
    return ''.join(str(e) for e in list)

#check total num 함수에는 [인원수, 인원수,...]형태로 파라미터가 들어간다고 가정(여기서는 ppl)
# ㅁ/ㅁ/ㅁ에서 추출한 ㅁ이 인원수로 설정되어서 [인원수, 인원수...]에 들어감
def check_total_num (ppl,n_ppl): 
    total_n = 0
    
    for i in range(len(ppl)):
        total_n += ppl[i]
    if total_n == n_ppl:
        return True   
    else :
        return False
            
#seat : string으로 묶은 빈자리
#num 검색할 자리수(인원수)
#빈자리가 존재하는 행 반환(!!!단, 행이 1부터 시작함)
def search_empty_seats(char,seats,num):
    count = 1
    for row in seats:
        if char*num in row:
            return count
        count +=1
    return False

def show_seats(seats,ppl,n_ppl) :
    new_seats = seats[:]

    for i in range(len(ppl)) :
        k = search_empty_seats('0',new_seats,ppl[i])
        each_string = new_seats[k-1]
        index = each_string.index('0'*ppl[i])
        each_string = list(new_seats[k-1])
        for m in range(ppl[i]) :
            each_string[index] = '*'
            index+=1
        new_seats[k-1] = list_value_into_string(each_string)
    print()
    return new_seats  
