# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'lastOrder.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import load_workbook
import time

libpaths = QtWidgets.QApplication.libraryPaths() #추가
libpaths.append("C:\\Users\사용자\AppData\Local\Programs\Python\Python37-32\Lib\site-packages\PyQt5\Qt\plugins")
QtWidgets.QApplication.setLibraryPaths(libpaths)

'''
from openpyxl import load_workbook
wb = load_workbook("상품관리.xlsx")
ws = wb.active
wl = load_workbook('결제로그.xlsx')
wls = wl.active
wm = load_workbook('회원관리.xlsx')
wms = wm.active
wbs = load_workbook("좌석관리.xlsx")
wss = wbs.active
'''
class Ui_MainWindow(object):
#------------------------------------------------------------------------------------------------------------------
    def check_id(self, id):
        from openpyxl import load_workbook
        wb = load_workbook("상품관리.xlsx")
        ws = wb.active
        wm = load_workbook('회원관리.xlsx')
        wms = wm.active
        
        for rowNum in range(2, ws.max_row+1): #1행은 헤더이므로 제외
            id_value = wms.cell(row=rowNum, column=2).value  #(B(2)열 : 아이디)
            if id_value == id :
                return rowNum
            else :
                continue
        return
    
    def bttnClicked(self,list):
        from openpyxl import load_workbook
        wb = load_workbook("상품관리.xlsx")
        ws = wb.active
        wl = load_workbook('결제로그.xlsx')
        wls = wl.active
        wm = load_workbook('회원관리.xlsx')
        wms = wm.active
        wbs = load_workbook("좌석관리.xlsx")
        wss = wbs.active
        
        #list : [user_id,charge_time,leftOver,payTime,payProduct,selectedSeat]
        now = time.localtime() #현재 날짜, 시간 저장
        s = "%02d월 %02d일 %02d시 %02d분 %02d초" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
        #print(s)
        wls.cell(row=wls.max_row,column=1).value = s
        wl.save('결제로그.xlsx')
        
        userRow = self.check_id(list[0])
        wms.cell(row=userRow, column=6).value+=list[1]
        wm.save('회원관리.xlsx') #충전 시간 저장
        #print('*')

        for rowNum in range(2, ws.max_row+1):
            ws.cell(row=rowNum,column=7).value=list[2][rowNum-2] #상품 재고 저장
        ws.cell(row=2,column=8).value+=list[3] #판매 금액 저장
        ws.cell(row=2,column=9).value+=list[4]
        wb.save('상품관리.xlsx')  
        #print('*')

        wss.cell(row=list[5]+1,column=3).value = True #선택 좌석 저장
        wbs.save('좌석관리.xlsx')
        #print('*')
        msgBox = QtWidgets.QMessageBox.about(None,"결제완료",'결제가 완료되었습니다')
        #sys.exit(app.exec_())
        #return
        
        
#------------------------------------------------------------------------------------------------------------------
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(344, 470)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(60, 50, 56, 12))
        self.label.setObjectName("label")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(60, 100, 71, 16))
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(60, 170, 101, 16))
        self.label_4.setObjectName("label_4")
        self.moneyCharge = QtWidgets.QLabel(self.centralwidget)
        self.moneyCharge.setGeometry(QtCore.QRect(210, 100, 56, 12))
        self.moneyCharge.setObjectName("moneyCharge")
        self.moneyItem = QtWidgets.QLabel(self.centralwidget)
        self.moneyItem.setGeometry(QtCore.QRect(210, 170, 56, 12))
        self.moneyItem.setObjectName("moneyItem")
        self.label_8 = QtWidgets.QLabel(self.centralwidget)
        self.label_8.setGeometry(QtCore.QRect(70, 310, 81, 16))
        self.label_8.setObjectName("label_8")
        self.moneyTotal = QtWidgets.QLabel(self.centralwidget)
        self.moneyTotal.setGeometry(QtCore.QRect(210, 310, 56, 12))
        self.moneyTotal.setObjectName("moneyTotal")
        self.bttnOrder = QtWidgets.QPushButton(self.centralwidget)
        self.bttnOrder.setGeometry(QtCore.QRect(204, 370, 71, 31))
        self.bttnOrder.setObjectName("bttnOrder")
        self.id = QtWidgets.QLabel(self.centralwidget)
        self.id.setGeometry(QtCore.QRect(130, 50, 130, 12))
        self.id.setObjectName("id")
        self.charge = QtWidgets.QLabel(self.centralwidget)
        self.charge.setGeometry(QtCore.QRect(80, 130, 56, 12))
        self.charge.setObjectName("charge")
        self.order = QtWidgets.QLabel(self.centralwidget)
        self.order.setGeometry(QtCore.QRect(80, 200, 81, 81))
        self.order.setObjectName("order")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 344, 17))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
#------------------------------------------------------------------------------------------------------------------
    #데이터 설정
        from openpyxl import load_workbook
        wb = load_workbook("상품관리.xlsx")
        ws = wb.active
        wl = load_workbook('결제로그.xlsx')
        wls = wl.active
        wm = load_workbook('회원관리.xlsx')
        wms = wm.active
        wbs = load_workbook("좌석관리.xlsx")
        wss = wbs.active
        
        selectedSeat = wls.cell(row = wls.max_row, column=7).value #선택좌석
        if selectedSeat == None:
            #wls.delete_rows(wls.max_row, 1)
            msgBox = QtWidgets.QMessageBox.about(None,"오류",'좌석 선택을 하지 않았습니다.\n좌석 선택을 하십시오.')
            #결제 창 종료
            sys.exit(app.exec_())
            return
        user_id = wls.cell(row = wls.max_row, column=2).value #사용자 아이디
        charge_time = wls.cell(row = wls.max_row, column=3).value #충전 시간
        if charge_time == None:
            charge_time = 0
        product_leftover = wls.cell(row = wls.max_row, column=4).value #상품재고
        payTime = wls.cell(row = wls.max_row, column=5).value #시간판매금액
        if payTime == None:
            payTime = 0
        payProduct = wls.cell(row = wls.max_row, column=6).value #상품판매금액
        if payProduct == None:
            payProduct = 0

        leftOver = product_leftover.split('/')
        
        menu = []
        for rowNum in range(2, ws.max_row+1):
            menu.append(ws.cell(row=rowNum,column=1).value)
            
        dict_bought={}
        
        i=0
        string=''
        for product in menu:
            left = ws.cell(row=i+2,column=7).value - int(leftOver[i])
            if left>0:
                string += menu[i]+' : '+str(left)+'개\n'
            leftOver[i] = int(leftOver[i])
            i+=1
            
        #print(dict_bought)    
        #print(leftOver)
        

        self.id.setText('%s' %user_id)
        self.charge.setText('%d시간' %charge_time)
        self.moneyCharge.setText('%d원' %payTime)
        self.moneyItem.setText('%d원' %payProduct)
        self.moneyTotal.setText('총 %d원' %(payTime+payProduct))
        
        self.order.setText(string)

        #print(user_id,charge_time,product_leftover,payTime,payProduct,selectedSeat)
        
        
#이벤트 헨들러
    #커넥션
        self.bttnOrder.clicked.connect(lambda: self.bttnClicked([user_id,charge_time,leftOver,payTime,payProduct,selectedSeat]))
        
#------------------------------------------------------------------------------------------------------------------

        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "결제"))
        self.label.setText(_translate("MainWindow", "아이디"))
        self.label_3.setText(_translate("MainWindow", "총 충전 시간"))
        self.label_4.setText(_translate("MainWindow", "상품 주문 내역"))
        self.moneyCharge.setText(_translate("MainWindow", "TextLabel"))
        self.moneyItem.setText(_translate("MainWindow", "TextLabel"))
        self.label_8.setText(_translate("MainWindow", "총 결제 금액"))
        self.moneyTotal.setText(_translate("MainWindow", "TextLabel"))
        self.bttnOrder.setText(_translate("MainWindow", "결제"))
        self.id.setText(_translate("MainWindow", "TextLabel"))
        self.charge.setText(_translate("MainWindow", "TextLabel"))
        self.order.setText(_translate("MainWindow", "TextLabel"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

