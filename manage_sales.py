import openpyxl

wb = openpyxl.load_workbook('상품관리.xlsx')
ws = wb['상품관리']

def manage_sales():
    total_timePrice = ws.cell(row=2, column=8).value
    total_productPrice = ws.cell(row=2, column=9).value

    sales = total_timePrice + total_productPrice
    print("총 시간 판매금액은 [%d]원 입니다."%total_timePrice)
    print("총 상품 판매금액은 [%d]원 입니다."%total_productPrice)
    print("총 합계 판매금액은 [%d]원 입니다."%sales)

#manage_sales()

    
    
    
